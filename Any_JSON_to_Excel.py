'''
Python script to read JSON files from an input directory path
Convert any type of nested JSON to a flat style JSON
Write the JSON data to an excel sheet
'''
import json
import codecs
#import itertools
import os
import sys
#from time import sleep
from datetime import datetime
import dateutil.parser
from openpyxl import Workbook
from flatten_dict import flatten
#import shutil
#from pprint import pprint

global isFirstRow
global row
global firstRowKey

def load_json_from_directories(rootPath, filterAssetCode):
    '''
    Checks if the file is .json extension or any other extension in a directory
    '''
    for root, dirs, files in os.walk(rootPath):
        for file in files:
            if os.path.splitext(file)[1] == '.json':
                print('Processing...'+file)
                load_json_line(os.path.join(root, file), filterAssetCode)
            else:
                print('Not a json file! Skipping...'+file)

def load_json_line(filename, filterAssetCode):
    '''
    Load the file and iterate line by line
    '''
    with codecs.open(filename, 'r') as data_file:
        for line in data_file:
            # LOAD EACH LINE FROM THE JSON FILE
            origDict = json.loads(line)
            flatDict = flatten(origDict, reducer='underscore')
            if filterAssetCode in (flatDict['Header_ID'], 'All'):
                write_to_excel(flatDict)

def write_to_excel(flatDict):
    '''
    Receive a JSON flat dictionary and fill it in the excel rows
    '''
    flatKey = list(flatDict.keys())

# FILLING THE FIRST ROW WITH COLUMN TITLES
    if isFirstRow:
        firstRowKey = flatKey
        col = 1
        for i in flatKey:
            workSheet.cell(row=row, column=col, value=i)
            col += 1
        row += 1
        isFirstRow = False

# FILLING THE ACTUAL VALUES FROM 2ND ROW ONWARDS
    if flatKey == firstRowKey:
        col = 1
        for j in flatKey:
            if j in ('Data_RTC', 'EventProcessedUtcTime', 'EventEnqueuedUtcTime'):
                cellValue = dateutil.parser.parse(flatDict[j])
            else:
                cellValue = flatDict[j]
            workSheet.cell(row=row, column=col, value=cellValue)
            col += 1
        row += 1

if __name__ == '__main__':
    sys.setrecursionlimit(1500)
    workBook = Workbook()
    workSheet = workBook.active
    workSheet.title = "JSON Output"
    isFirstRow = True
    row = 1

    rootPath = input('Enter full path of the raw data parent directory: ')

    isFilterField = input('Do you want to filter specific Asset code (Y/N) ? ')

    if isFilterField.lower() == 'y':
        filterAssetCode = input('Enter the 8 digit asset code: ')
    else:
        filterAssetCode = 'All'
    load_json_from_directories(rootPath, filterAssetCode)

#SAVE THE EXCEL FILE
    outputFileName = filterAssetCode + '_' + datetime.now().strftime("%Y%m%d%H%M%S") + '.xlsx'
    print('Saving the excel file '+outputFileName)
    workBook.save(os.path.join(rootPath, outputFileName))
    print('------------------------------Completed Successfully------------------------------')
    #sleep(0.2)
    #LAUNCH THE EXCEL FILE
    #print('Opening the excel file '+outputFileName)
    #os.startfile(os.path.join(rootPath, outputFileName))
