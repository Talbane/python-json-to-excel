'''
FULL APPLICATION VERSION 0.5
'''
import json
import codecs
import itertools
import os
import sys
from time import sleep
from datetime import datetime
import dateutil.parser
from openpyxl import Workbook
#import shutil
#from pprint import pprint

def load_json_from_directories(root_path, cur_path, asset_code):
    for root,dirs,files in os.walk(root_path):
        for file in files:
            if os.path.splitext(file)[1] == '.json':
                print('Processing...'+file)
                load_json_line(os.path.join(root,file),asset_code)
            else:
                print('Not a json file! Skipping...'+file)

def load_json_line(filename, asset_code):
    with codecs.open(filename, 'r') as data_file:
        for line in data_file:
            # LOAD EACH LINE FROM THE JSON FILE
            data_item = json.loads(line)
            if data_item['Header']['ID'] == asset_code:
            	write_to_excel(data_item)
            elif asset_code == 'All':
            	write_to_excel(data_item)

def write_to_excel(data_item):
    global FIRST_ROW
    global row
    datakeyfull =list(data_item['Data'].keys())
    tempindex = datakeyfull.index('DG1')
    # FILLING THE FIRST ROW WITH COLUMN TITLES
    if FIRST_ROW == True:
        headerkey = list(data_item['Header'].keys())
        datakey = datakeyfull[:tempindex]
        dg1key=["DG1-"+ el for el in list(data_item['Data']['DG1'].keys())]
        dg2key=["DG2-"+ el for el in list(data_item['Data']['DG2'].keys())]
        dg3key=["DG3-"+ el for el in list(data_item['Data']['DG3'].keys())]
        dg4key=["DG4-"+ el for el in list(data_item['Data']['DG4'].keys())]
        dg5key=["DG5-"+ el for el in list(data_item['Data']['DG5'].keys())]
        dg6key=["DG6-"+ el for el in list(data_item['Data']['DG6'].keys())]

        col = 1
        for i in itertools.chain(headerkey,datakey,dg1key,dg2key,dg3key,dg4key,dg5key,dg6key):
            WS1.cell(row=row, column=col, value=i)
            col += 1
        WS1.cell(row=row, column=col, value='EventProcessedUtcTime')
        WS1.cell(row=row, column=col+1, value='EventEnqueuedUtcTime')
        row += 1
        FIRST_ROW = False

    headervalue = list(data_item['Header'].values())
    datavaluefull = list(data_item['Data'].values())
    datavalue = datavaluefull[:tempindex]
    dg1value = list(data_item['Data']['DG1'].values())
    dg2value = list(data_item['Data']['DG2'].values())
    dg3value = list(data_item['Data']['DG3'].values())
    dg4value = list(data_item['Data']['DG4'].values())
    dg5value = list(data_item['Data']['DG5'].values())
    dg6value = list(data_item['Data']['DG6'].values())


    # FILLING THE ACTUAL VALUES FROM 2ND ROW ONWARDS
    col = 1
    for j in itertools.chain(headervalue,datavalue,dg1value,dg2value,dg3value,dg4value,dg5value,dg6value):
        if col == len(headervalue)+1:
            rtcdatetime = dateutil.parser.parse(j)
            WS1.cell(row=row, column=col, value=rtcdatetime)
            col += 1
        else:
            WS1.cell(row=row, column=col, value=j)
            col += 1
    WS1.cell(row=row, column=col, value=dateutil.parser.parse(data_item['EventProcessedUtcTime']))
    WS1.cell(row=row, column=col+1, value=dateutil.parser.parse(data_item['EventEnqueuedUtcTime']))
    row += 1

if __name__ == '__main__':
	sys.setrecursionlimit(1500)
	WB = Workbook()
	WS1 = WB.active
	WS1.title = "JSON Output"
	FIRST_ROW = True
	row = 1

	root_path = input('Enter full path of the raw data parent directory: ')
	cur_path = root_path

	FILTER_AC = input('Do you want to filter specific Asset code (Y/N) ? ')

	if FILTER_AC.lower() == 'y':
	    asset_code = input('Enter the 8 digit asset code: ')
	else:
	    asset_code = 'All'
	load_json_from_directories(root_path, cur_path, asset_code)
	print('Completed Successfully')
	#SAVE THE EXCEL FILE
	OUTPUT_FILENAME = asset_code + '_' + datetime.now().strftime("%Y%m%d%H%M%S") + '.xlsx'
	print('Saving the excel file '+OUTPUT_FILENAME)
	WB.save(os.path.join(root_path, OUTPUT_FILENAME))
	#sleep(0.2)
	#LAUNCH THE EXCEL FILE
	#print('Opening the excel file '+OUTPUT_FILENAME)
	#os.startfile(os.path.join(root_path, OUTPUT_FILENAME))
    