'''
FULL APPLICATION VERSION 0.5 WITH OS.WALK
Concrete pump parsing
'''
import json
import codecs
import itertools
import os
import sys
from time import sleep
from datetime import datetime
import dateutil.parser
from openpyxl import Workbook
#import shutil
#from pprint import pprint

def load_json_from_directories(root_path, cur_path, asset_code):
    for root,dirs,files in os.walk(root_path):
        for file in files:
            if os.path.splitext(file)[1] == '.json':
                print('Processing...'+file)
                load_json_line(os.path.join(root,file),asset_code)
            else:
                print('Not a json file! Skipping...'+file)

def load_json_line(filename, asset_code):
    with codecs.open(filename, 'r') as data_file:
        for line in data_file:
            # LOAD EACH LINE FROM THE JSON FILE
            data_item = json.loads(line)
            if data_item['AssetId'] == asset_code:
            	write_to_excel(data_item)
            elif asset_code == 'All':
            	write_to_excel(data_item)

def write_to_excel(data_item):
    global FIRST_ROW
    global row
    gatewaycsvvalue = list(data_item['KeyValues'].split(sep=','))
    serverdatetimevalue = dateutil.parser.parse(data_item['ServerDateTime'])
    #headervalue = list(data_item['Header'].values())
    #datavalue = list(data_item['Data'].values())
    # FILLING THE FIRST ROW WITH COLUMN TITLES
    if FIRST_ROW == True:
        titlekey = ['ID','Type','Mvmt','Fuel','SNo','IMEI','CCode','Version','Volt','Batt','Lat','Long','RTC','HOP','HOT','EOP','EOT','Strokes','Volume','Ign','Eng','Enghrs','LoadType','FuelLevel','RPM','ErrorCode']
        #headerkey = list(data_item['Header'].keys())
        #datakey = list(data_item['Data'].keys())
        col = 1
        #for i in itertools.chain(headerkey, datakey):
        for key in titlekey:
            WS1.cell(row=row, column=col, value=key)
            col += 1
        WS1.cell(row=row, column=col, value='ServerDateTime')
        #WS1.cell(row=row, column=col+1, value='sadatetimeist')
        row += 1
        FIRST_ROW = False

    # FILLING THE ACTUAL VALUES FROM 2ND ROW ONWARDS
    col = 1
    for value in gatewaycsvvalue:
        if col == 13:
            #rtcdatetime = dateutil.parser.parse(value)
            WS1.cell(row=row, column=col, value=value)
            col += 1
        else:
            WS1.cell(row=row, column=col, value=value)
            col += 1
    WS1.cell(row=row, column=col, value=serverdatetimevalue)
    #WS1.cell(row=row, column=col+1, value=dateutil.parser.parse(data_item['sadatetimeist']))
    row += 1

if __name__ == '__main__':
	sys.setrecursionlimit(1500)
	WB = Workbook()
	WS1 = WB.active
	WS1.title = "JSON Output"
	FIRST_ROW = True
	row = 1

	root_path = input('Enter full path of the raw data parent directory: ')
	cur_path = root_path

	FILTER_AC = input('Do you want to filter specific Asset code (Y/N) ? ')

	if FILTER_AC.lower() == 'y':
	    asset_code = input('Enter the 8 digit asset code: ')
	else:
	    asset_code = 'All'
	load_json_from_directories(root_path, cur_path, asset_code)
	print('Completed Successfully')
	#SAVE THE EXCEL FILE
	OUTPUT_FILENAME = datetime.now().strftime("%Y%m%d%H%M%S") + '.xlsx'
	print('Saving the excel file '+OUTPUT_FILENAME)
	WB.save(os.path.join(root_path, OUTPUT_FILENAME))
	#sleep(0.2)
	#LAUNCH THE EXCEL FILE
	#print('Opening the excel file '+OUTPUT_FILENAME)
	#os.startfile(os.path.join(root_path, OUTPUT_FILENAME))
    